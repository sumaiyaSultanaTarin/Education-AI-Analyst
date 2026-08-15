"""XLSX row extraction, used by the Document Ingestion Agent."""

from openpyxl import load_workbook


def extract_xlsx_content(path: str) -> dict:
    """Extract every sheet as a list of row-dicts keyed by its header row.

    Returns {"sheets": {sheet_name: list[dict]}}.
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            sheets[sheet_name] = []
            continue

        sheets[sheet_name] = [
            dict(zip(header, row))
            for row in rows
            if any(cell is not None for cell in row)
        ]

    return {"sheets": sheets}
